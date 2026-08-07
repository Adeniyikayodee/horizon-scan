"""F5: "verified" meant two different things and the workbook printed one word.

Either a quote was fetched from the cited source and found in it, or nothing could
be fetched and the status rests on the model's word. Across the archived runs the
second case was the majority: reading_grounded was None on 29 of 29 rows in
a158691ae1, and quote_grounded was never True in several runs.
"""
from __future__ import annotations

import glob
import json

import pytest
from openpyxl import load_workbook

from scan import config, guardrail, io_xlsx


def _row(status="verified", qg=None, rg=None, **kw):
    return {"name": "A program", "org": "Org", "verification": {"status": status},
            "quote_grounded": qg, "reading_grounded": rg, **kw}


# --- the label itself ---------------------------------------------------------------
def test_verified_with_a_grounded_quote_is_checked():
    assert guardrail.evidence_check(_row(qg=True)) == "checked"


def test_verified_with_a_grounded_reading_is_checked():
    assert guardrail.evidence_check(_row(rg=True)) == "checked"


def test_verified_with_nothing_fetchable_is_unchecked():
    assert guardrail.evidence_check(_row(qg=None, rg=None)) == "unchecked"


def test_partial_rows_carry_no_evidence_check():
    assert guardrail.evidence_check(_row(status="partial", qg=True)) == ""


# --- the status vocabulary must not move --------------------------------------------
def test_verification_status_is_untouched():
    """Two consumers parse this string, the stage-two rejoin and the app counter.
    The firmness rides in its own column precisely so this one cannot move."""
    r = _row(qg=True)
    guardrail.evidence_check(r)
    assert r["verification"]["status"] == "verified"


def test_longlist_keeps_a_machine_readable_verification_column(tmp_path, monkeypatch):
    monkeypatch.setattr(config, "REVIEW_DIR", tmp_path)
    monkeypatch.setattr(config, "WORK_DIR", tmp_path)
    io_xlsx.write_longlist([_row(qg=True), _row(status="partial")])
    ws = load_workbook(tmp_path / "longlist.xlsx").active
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c).lower() for c in rows[0]]
    assert "verification" in header and "evidence check" in header
    vi, ei = header.index("verification"), header.index("evidence check")
    assert rows[1][vi] == "verified" and rows[1][ei] == "checked"
    # a partial row leaves the cell empty, which openpyxl reads back as None
    assert rows[2][vi] == "partial" and not rows[2][ei]


# --- the rejoin still works, which is the regression that matters --------------------
def test_rejoin_survives_the_new_column(tmp_path, monkeypatch):
    monkeypatch.setattr(config, "REVIEW_DIR", tmp_path)
    monkeypatch.setattr(config, "WORK_DIR", tmp_path)
    rows = [_row(qg=True, what="What it is", url="https://example.org/a", quotes=["q"]),
            _row(status="partial", name="Another program")]
    io_xlsx.write_longlist(rows)
    kept = io_xlsx.read_kept_longlist()
    assert len(kept) == 2
    assert [k["verification"]["status"] for k in kept] == ["verified", "partial"]
    assert kept[0]["quotes"] == ["q"], "the full row must still rejoin by rid"
    assert kept[0]["what"] == "What it is"


def test_map_carries_the_evidence_check(tmp_path, monkeypatch):
    monkeypatch.setattr(config, "OUT_DIR", tmp_path)
    io_xlsx.write_map([_row(qg=True)], [{"name": "A theme", "members": ["A program"]}])
    ws = load_workbook(tmp_path / "innovation_map.xlsx").active
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c).lower() for c in rows[0]]
    assert header.index("evidence check") == header.index("verification") + 1
    assert rows[1][header.index("evidence check")] == "checked"


# --- replay: how much of the archive was actually checked ----------------------------
def test_replay_labels_the_archive_without_crashing():
    checked = unchecked = partial = 0
    for f in glob.glob("runs/*/work/longlist_full.json"):
        try:
            rows = json.loads(open(f, encoding="utf-8").read())
        except Exception:
            continue
        for r in rows:
            got = guardrail.evidence_check(r)
            assert got in ("checked", "unchecked", "")
            checked += got == "checked"
            unchecked += got == "unchecked"
            partial += got == ""
    if checked + unchecked:
        # the whole point of the split: the archive is not uniformly "verified"
        assert unchecked > 0, "no unchecked rows found, the split would be pointless"
