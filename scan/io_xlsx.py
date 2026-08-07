"""Excel and markdown I/O: read the org sheet, write the review longlist and
the stage-2 deliverables. Uses openpyxl directly to control author metadata.
"""
from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from . import config, guardrail

AUTHOR = "Kayode Adeniyi"

# descriptor words dropped before matching, so trivial variants of the same
# organization collapse while distinct organizations do not
_ORG_STOP = {"the", "and", "of", "for", "a", "an", "group", "ltd", "limited", "inc",
             "llc", "foundation", "fund", "institute", "initiative", "organization",
             "organisation", "org"}


def _norm_org(name: str) -> str:
    """The normalized full-name key: lowercased, parentheticals dropped, punctuation
    flattened, and descriptor words removed. If a name is made ENTIRELY of descriptor
    words ('The Fund'), the stripped form falls back to the full token list so the
    organization is never reduced to an empty key and silently dropped."""
    s = re.sub(r"\([^)]*\)", " ", (name or "").lower())     # drop "(AERC)" etc.
    s = re.sub(r"[^a-z0-9]+", " ", s)
    toks = s.split()
    kept = [t for t in toks if t not in _ORG_STOP]
    return " ".join(kept or toks).strip()


# Articles only. A PROGRAM name is not an organization name: "Fund", "Institute",
# "Initiative", and "Foundation" are the words that tell two real programs apart,
# so the organization stop list must never be used on one.
_APPROACH_STOP = {"the", "and", "of", "for", "a", "an"}


def _norm_approach(name: str) -> str:
    """The matching key for a PROGRAM name. Parentheticals and punctuation are
    dropped, so "Blue Economy Program (PROFISHBLUE)" still collapses onto "Blue
    Economy Program", but every descriptor word is kept, so "Blue Economy Fund" and
    "Blue Economy Initiative" stay two programs. Deliberately not _norm_org: that
    one strips exactly the words this one needs."""
    s = re.sub(r"\([^)]*\)", " ", (name or "").lower())
    s = re.sub(r"[^a-z0-9]+", " ", s)
    toks = [t for t in s.split() if t not in _APPROACH_STOP]
    return " ".join(toks).strip()


def _org_sig(name: str) -> dict:
    """A match signature: the normalized full name, the acronym the organization
    DECLARES in parentheses, the acronym DERIVED from its initials, and whether the
    name is itself a bare acronym. Keeping declared and derived apart lets 'AfDB'
    match 'African Development Bank (AfDB)' and 'ADB' match 'Asian Development Bank
    (ADB)' without either bare acronym colliding with the other bank."""
    low = (name or "").lower()
    declared = {re.sub(r"[^a-z0-9]+", "", p) for p in re.findall(r"\(([^)]*)\)", low)}
    declared = {a for a in declared if 2 <= len(a) <= 8}
    base = _norm_org(name)
    words = base.split()
    derived = {"".join(w[0] for w in words)} if len(words) >= 2 else set()
    is_acro = len(words) == 1 and base.isalpha() and 2 <= len(base) <= 8
    return {"base": base, "declared": declared, "derived": derived, "is_acro": is_acro}


def _acro_hits(a: dict, b: dict) -> bool:
    """Does bare-acronym `a` name full-name `b`? Only when `a` equals `b`'s declared
    acronym, or `b` declares none and `a` equals `b`'s initials."""
    if not a["is_acro"]:
        return False
    return a["base"] in b["declared"] or (not b["declared"] and a["base"] in b["derived"])


def _same_org(a: dict, b: dict) -> bool:
    if not a["base"] or not b["base"]:
        return False
    if a["base"] == b["base"]:                              # same full name (or same acronym)
        return True
    return _acro_hits(a, b) or _acro_hits(b, a)


def merge_orgs(primary: list[dict], additions: list[dict]) -> list[dict]:
    """Join two organization lists into one deduped roster with provenance. Order is
    stable: the primary list first, then analyst-only additions appended. On a match
    the entry is kept once, marked as coming from both when the sides differ, and any
    field the entry was missing is filled from the other, so no analyst detail is
    lost. Every returned org carries a `source` (discovered, analyst, or both) and
    ids are reassigned contiguously."""
    out: list[dict] = []
    sigs: list[dict] = []

    def absorb(o: dict, src: str) -> None:
        sig = _org_sig(o.get("name", ""))
        if not sig["base"]:
            return
        for i, existing in enumerate(sigs):
            if _same_org(sig, existing):
                e = out[i]
                for f in ("type", "region", "why"):
                    if not (e.get(f) or "").strip() and (o.get(f) or "").strip():
                        e[f] = o[f]
                if e["source"] != src:
                    e["source"] = "both"
                existing["declared"] |= sig["declared"]     # remember both spellings
                existing["derived"] |= sig["derived"]
                return
        m = dict(o)
        m["source"] = (o.get("source") or "").strip() or src   # keep an existing label
        if m["source"] == "analyst" and not (m.get("why") or "").strip():
            m["why"] = "Added by the analyst."
        out.append(m)
        sigs.append(sig)

    for o in primary:
        absorb(o, "discovered")
    for o in additions:
        absorb(o, "analyst")
    for i, o in enumerate(out, start=1):
        o["id"] = f"O{i:03d}"
    return out


def _stamp(wb: Workbook) -> None:
    wb.properties.creator = AUTHOR
    wb.properties.lastModifiedBy = AUTHOR


# --- input ---
def read_orgs(path: Path | None = None) -> list[dict[str, str]]:
    path = path or config.ORG_SHEET
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    orgs = []
    for i, r in enumerate(rows[1:], start=1):
        d = {header[j]: ("" if v is None else str(v)) for j, v in enumerate(r) if j < len(header)}
        if not d.get("name"):
            continue
        d.setdefault("id", f"O{i:03d}")
        orgs.append(d)
    return orgs


def write_orgs_list(orgs: list[dict], path: Path | None = None) -> Path:
    """Write a discovered organization list to xlsx (editable by the researcher)."""
    path = path or config.ORG_SHEET
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "organizations"
    ws.append(["id", "name", "type", "region", "why", "source"])
    for i, o in enumerate(orgs, start=1):
        ws.append([f"O{i:03d}", o.get("name", ""), o.get("type", ""), o.get("region", ""),
                   o.get("why", ""), o.get("source", "")])
    _stamp(wb)
    wb.save(path)
    return path


def write_sample_orgs(path: Path | None = None) -> Path:
    path = path or config.ORG_SHEET
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "organizations"
    ws.append(["id", "name", "type", "region", "status"])
    sample = [
        ("O001", "African Economic Research Consortium (AERC)", "African policy institute", "Africa", ""),
        ("O002", "Policy Center for the New South", "African policy institute", "Africa", ""),
        ("O003", "UN Trade and Development (UNCTAD)", "Multilateral", "Global", ""),
    ]
    for row in sample:
        ws.append(row)
    _stamp(wb)
    wb.save(path)
    return path


# --- review (stage 1 out) ---
def write_longlist(rows: list[dict[str, Any]]) -> Path:
    path = config.REVIEW_DIR / "longlist.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "longlist"
    crit = config.active_spec().get("criteria", [])
    cols = (["rid", "keep", "org", "approach", "band", "what", "year"]
            + [c["name"] for c in crit]
            + ["overall", "verification", "evidence check", "source"])
    ws.append(cols)
    for i, r in enumerate(rows, start=1):
        r.setdefault("rid", f"R{i:04d}")                 # stable id for the stage-2 rejoin
        s = r.get("score", {}) or {}
        v = r.get("verification", {}) or {}
        row = [r["rid"], "Y", r.get("org", ""), r.get("name", ""), r.get("band", ""),
               r.get("what", ""), r.get("year", "")]
        row += [s.get(c["key"], "") for c in crit]
        row += [s.get("overall", ""), v.get("status", ""),
                guardrail.evidence_check(r), r.get("url", "")]
        ws.append(row)
    _stamp(wb)
    wb.save(path)
    # persist the FULL rows so stage two can rejoin evidence, quotes, and scores that
    # do not fit the review sheet, keyed by the same rid
    (config.WORK_DIR / "longlist_full.json").write_text(
        json.dumps(rows, ensure_ascii=False), encoding="utf-8")
    return path


def _full_rows_by_rid() -> dict[str, dict]:
    p = config.WORK_DIR / "longlist_full.json"
    if not p.exists():
        return {}
    return {r.get("rid", ""): r for r in json.loads(p.read_text(encoding="utf-8"))}


def read_kept_longlist() -> list[dict[str, Any]]:
    """The kept rows for stage two. The human keep decision and any cell edits come
    from the review workbook, but the full evidence, quotes, and score marks are
    rejoined from the persisted stage-one rows by their stable rid, so the memo and
    the theming see the whole record rather than the thin review view."""
    path = config.REVIEW_DIR / "longlist.xlsx"
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    full_by_rid = _full_rows_by_rid()
    kept: list[dict[str, Any]] = []
    for r in rows[1:]:
        d = {header[j]: v for j, v in enumerate(r) if j < len(header)}
        if str(d.get("keep", "")).strip().upper() != "Y":
            continue
        base = dict(full_by_rid.get(str(d.get("rid", "") or ""), {}))
        # overlay the analyst's cell edits from the review sheet onto the full row,
        # so human corrections win while evidence, quotes, and scores are preserved
        row = {**base,
               "org": d.get("org", "") or base.get("org", ""),
               "name": d.get("approach", "") or base.get("name", ""),
               "band": d.get("band", "") or base.get("band", ""),
               "what": d.get("what", "") or base.get("what", ""),
               "overall": d.get("overall", "") or base.get("overall", ""),
               "year": d.get("year", "") or base.get("year", ""),
               "url": d.get("source", "") or base.get("url", "")}
        vstatus = str(d.get("verification", "") or "").strip()
        row["verification"] = {**(base.get("verification") or {}),
                               **({"status": vstatus} if vstatus else {})}
        kept.append(row)
    return kept


def write_open_questions(rows: list[dict[str, Any]], dropped: list[dict[str, Any]]) -> Path:
    path = config.REVIEW_DIR / "open_questions.md"
    lines = ["# Open questions\n"]
    partial = [r for r in rows if (r.get("verification", {}) or {}).get("status") == "partial"]
    lines.append("## Partial, needs a primary\n")
    for r in partial:
        note = (r.get("verification", {}) or {}).get("note", "")
        lines.append(f"- {r.get('org','')}: {r.get('name','')}, {note}")
    lines.append("\n## Set aside, and why\n")
    lines.append("Every candidate that did not make the longlist, with the stage it went at "
                 "and the reason. A drop here is never seen again, so it is worth a look when "
                 "the list is long.\n")
    for d in dropped:
        stage = d.get("stage", "") or "not kept at reading"
        reason = d.get("reason", "") or "no reason given"
        lines.append(f"- {d.get('org','')}: {d.get('name','')} [{stage}] {reason}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path


def write_theme_screen(themes: list[dict[str, Any]]) -> Path:
    """What the existing-portfolio screen did, so the analyst can see and argue with
    it. A theme held back here is not a failure of the scan, it is the scan refusing
    to recommend work the institute already runs. Edit the spec's excluded_areas to
    change the call, never the code."""
    path = config.REVIEW_DIR / "theme_screen.md"
    held = [t for t in themes if t.get("screened")]
    near = [t for t in themes if t.get("screen_note") and not t.get("screened")]
    lines = ["# Existing-portfolio screen\n",
             f"{len(held)} theme(s) held back to existing work, {len(near)} near the line.\n",
             "## Held back to existing work, posture deepen\n"]
    lines += [f"- {t.get('name','')}: {t['screened']}" for t in held] or ["- none"]
    lines += ["\n## Near the line, left as the model tagged them\n"]
    lines += [f"- {t.get('name','')} [{t.get('tag','')}, {t.get('posture','')}]: {t['screen_note']}"
              for t in near] or ["- none"]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path


def write_hunches(patterns: list[dict[str, str]]) -> Path:
    path = config.REVIEW_DIR / "hunches.md"
    lines = [
        "# Hunches\n",
        "Seeded from cross-org patterns. Overwrite with your own reading, this",
        "is the part a tool misses and it feeds the theming in stage 2.\n",
    ]
    for p in patterns:
        lines.append(f"- {p.get('name','')}: {p.get('note','')}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path


# --- deliverables (stage 2 out) ---
def write_policy_violations(items: list[tuple[str, str, list[str]]]) -> Path | None:
    """Record any policy violation left after the mechanical scrub, next to the
    delivered files rather than instead of them. Returns None when everything is
    clean, and removes a stale file from an earlier run so the absence of the file
    always means a clean run."""
    path = config.REVIEW_DIR / "policy_violations.md"
    flagged = [(label, text, hits) for label, text, hits in items if hits]
    if not flagged:
        path.unlink(missing_ok=True)
        return None
    lines = ["# Policy violations\n",
             "These were found in the delivered files. The files were written anyway, "
             "so nothing is lost, but each line below needs rewriting by hand before "
             "the work goes out. A machine must not fix these: removing the words and "
             "keeping the meaning are two different jobs.\n"]
    for label, text, hits in flagged:
        lines.append(f"## {label}\n")
        for h in hits:
            lines.append(f"- **{h}**")
            ctx = guardrail.context_for(text, h)
            if ctx:
                lines.append(f"  > {ctx}")
        lines.append("")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path


def write_scorecard(themes: list[dict[str, Any]], intro: str) -> Path:
    path = config.OUT_DIR / "theme_scorecard.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Theme scorecard"
    ws.append([intro])
    ws.append([])
    crit = config.active_spec().get("criteria", [])
    cols = ["Theme", "Standing"] + [c["name"] for c in crit] + ["Posture", "Evidence", "Top area", "Marquee"]
    ws.append(cols)
    for t in themes:
        row = [t.get("name", ""), t.get("tag", "")] + [t.get(c["key"], "") for c in crit]
        row += [t.get("posture", ""), t.get("evidence", ""), "yes" if t.get("top2") else "", t.get("marquee", "")]
        ws.append(row)
    _stamp(wb)
    wb.save(path)
    return path


def write_map(rows: list[dict[str, Any]], themes: list[dict[str, Any]]) -> Path:
    path = config.OUT_DIR / "innovation_map.xlsx"
    member_theme = {}
    for t in themes:
        for m in t.get("members", []):
            member_theme[m] = t["name"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Scan"
    ws.append(["Theme", "Approach", "Band", "What", "Year", "Overall", "Verification",
               "Evidence check", "Source"])
    for r in rows:
        ws.append([
            member_theme.get(r.get("name", ""), ""), r.get("name", ""), r.get("band", ""),
            r.get("what", ""), r.get("year", ""), r.get("overall", ""),
            (r.get("verification", {}) or {}).get("status", ""),
            guardrail.evidence_check(r), r.get("url", ""),
        ])
    _stamp(wb)
    wb.save(path)
    return path


def write_memo(markdown: str) -> Path:
    # Not gated here. The policy check runs in run_stage2 BEFORE any write, so a
    # violation produces a note beside the deliverables instead of an exception that
    # destroys a run already paid for. See guardrail.finalize.
    path = config.OUT_DIR / "synthesis_memo.md"
    path.write_text(markdown.rstrip() + "\n", encoding="utf-8")
    return path
